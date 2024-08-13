#! python3
# multiplication_table - Generate a multiplication table in an excel file using Python

import openpyxl, sys, os
from openpyxl.styles import Font

num = None

while True:
    if len(sys.argv) == 2:
        num = int(sys.argv[1])
        break
    else:
        print("Enter a multiplication number: python file_name.py number")

if os.path.exists("multiplication_table.xlsx"):
    wb = openpyxl.load_workbook("multiplication_table.xlsx")
else:
    wb = openpyxl.Workbook()

# Use the active sheet instead of accessing by name
sheet = wb.active


# Check if the sheet's title is 'Sheet' and rename it if necessary
if sheet.title == "Sheet":
    sheet.title = "Multiplication Table"


head_font = Font(size=12, bold=True)
sheet['A1'] = 'Multiplication Table'
sheet['A1'].font = head_font


i = 1
for no in range(2, num + 2):
    sheet.cell(row=no, column=1).value = i
    sheet.cell(row=no, column=1).font = head_font
    sheet.cell(row=1, column=no).value = i
    sheet.cell(row=1, column=no).font = head_font
    for colNum in range(2, num + 2):
        sheet.cell(row=no, column=colNum).value = i * (colNum-1)
    i+=1


wb.save("multiplication_table.xlsx")
