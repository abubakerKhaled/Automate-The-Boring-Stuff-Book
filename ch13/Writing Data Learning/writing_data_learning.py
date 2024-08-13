import openpyxl

# Creating a new Excel workbook
wb = openpyxl.Workbook()

# Print the default sheet names in the workbook (usually, there's one default sheet)
print("Initial sheet names:", wb.sheetnames)

# Access the active sheet (the sheet that is currently open)
sheet = wb.active
print("Active sheet title:", sheet.title)

# Rename the active sheet
sheet.title = "Spam Bacon Eggs Sheet"
print("Renamed active sheet title:", sheet.title)

# Create a new sheet at the first index (position 0) with a custom title
wb.create_sheet(title="First Index", index=0)
print("Sheet names after adding a new sheet at index 0:", wb.sheetnames)

# Create a new sheet at the end (default behavior when no index is specified)
wb.create_sheet()
print("Sheet names after adding a new sheet at the end:", wb.sheetnames)

# Delete the sheet named 'Sheet' (the default sheet if it still exists)
del wb["Sheet"]
print("Sheet names after deleting the default 'Sheet':", wb.sheetnames)

# Create a new sheet again with the default name 'Sheet'
wb.create_sheet()
sheet = wb["Sheet"]

# Write a value to cell A1 in the newly created sheet
sheet["A1"] = "Hello, Excel"
print("Value in cell A1 of the new 'Sheet':", sheet["A1"].value)

# Save the workbook to a file named 'writing_learning.xlsx'
wb.save("writing_learning.xlsx")

# Additional exploration:

# 1. Changing the tab color of a sheet
sheet.title = "Colored Tab Sheet"
sheet.sheet_properties.tabColor = "1072BA"  # Set tab color to a shade of blue
print(f"The tab color of '{sheet.title}' sheet has been changed.")

# 2. Adding a formula to a cell
sheet["B1"] = "=SUM(A1:A10)"
print(f"Formula in cell B1: {sheet['B1'].value}")

# 3. Merging and unmerging cells
sheet.merge_cells("C1:D1")  # Merge cells from C1 to D1
sheet["C1"] = "Merged Cells"
print(f"Value in merged cells C1:D1: {sheet['C1'].value}")

sheet.unmerge_cells("C1:D1")  # Unmerge the cells
print("Cells C1:D1 have been unmerged.")

# 4. Adjusting column width and row height
sheet.column_dimensions["A"].width = 20  # Set the width of column A
sheet.row_dimensions[1].height = 40  # Set the height of row 1
print("Column width and row height adjusted for column A and row 1.")

# 5. Freezing panes
sheet.freeze_panes = (
    "B2"  # Freeze the pane at B2 (top row and first column will be frozen)
)
print(f"Pane frozen at {sheet.freeze_panes}.")

# Save changes to the workbook
wb.save("writing_learning_extended.xlsx")
