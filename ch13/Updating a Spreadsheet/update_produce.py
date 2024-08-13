# 1 python3
# update_produce - Corrects costs in produce sales spreadsheet.

import openpyxl

print('Loading Workbook...')
wb = openpyxl.load_workbook("produceSales.xlsx")

wb['Sheet'].title = "ProduceSales" 
sheet = wb['ProduceSales']

# The produce types and their updated prices
PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# TODO: Loop through the rows and update the prices.
for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row = row_num, column = 1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row = row_num, column = 2).value = PRICE_UPDATES[produce_name]


wb.save("updatedProduceSales.xlsx")
