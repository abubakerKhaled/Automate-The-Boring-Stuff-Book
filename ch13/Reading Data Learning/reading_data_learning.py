import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Load the workbook from the file 'example.xlsx'
wb = openpyxl.load_workbook("example.xlsx")

# Print all sheet names in the workbook
print("Sheet names in the workbook:", wb.sheetnames)

# Access a specific sheet by its name
sheet = wb["Sheet3"]
print("Selected sheet title:", sheet.title)

# Access the active sheet (the sheet that was last opened)
another_sheet = wb.active
print("Active sheet title:", another_sheet.title)

# Print the minimum and maximum row and column indices
print("Minimum row:", another_sheet.min_row)
print("Maximum row:", another_sheet.max_row)
print("Minimum column:", another_sheet.min_column)
print("Maximum column:", another_sheet.max_column)

# Access and print the value of specific cells in the active sheet
print("Value of cell A1:", another_sheet["A1"].value)
print("Type of value in cell A1:", type(another_sheet["A1"].value))
print("Value of cell B1:", another_sheet["B1"].value)

# Store the value of cell B1 in a variable
value = another_sheet["B1"].value

# Access and print cell values using row and column indices
print("Value in row 1, column 1 (A1):", another_sheet.cell(row=1, column=1).value)
print("Value in row 1, column 2 (B1):", another_sheet.cell(row=1, column=2).value)
print("Value in row 1, column 3 (C1):", another_sheet.cell(row=1, column=3).value)

# Get and print the column letter corresponding to a column index
print("Column letter for column 1:", get_column_letter(1))
print("Column letter for column 2:", get_column_letter(2))
print("Column letter for column 27:", get_column_letter(27))
print("Column letter for column 999:", get_column_letter(999))

# Print the column letter of the last column in the active sheet
print("Column letter of the last column:", get_column_letter(another_sheet.max_column))

# Print the values in a range of cells (A1 to C3)
print("Values in the range A1:C3:")
for row_of_cell_objects in another_sheet["A1":"C3"]:
    for cell_object in row_of_cell_objects:
        # Print the cell's coordinate and its value
        print(f"Cell {cell_object.coordinate}: {cell_object.value}")
    print("--- END OF ROW ---")

